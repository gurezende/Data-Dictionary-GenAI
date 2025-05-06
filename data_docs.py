# Imports
import os
import json
import time
import tempfile
import streamlit as st
from textwrap import dedent

from agno.agent import Agent
from agno.models.google import Gemini
from agno.tools.file import FileTools

from openpyxl import load_workbook
from openpyxl.comments import Comment
import pandas as pd


def convert_to_csv(file_path:str):
   """
    Use this tool to convert the excel file to CSV.

    * file_path: Path to the Excel file to be converted
    """
   # Load the file  
   df = pd.read_excel(file_path).head(10)

   # Convert to CSV
   st.write("Converting to CSV... :leftwards_arrow_with_hook:")
   return df.to_csv('temp.csv', index=False)


# Custom Tool to add comments to the header of an Excel file
def add_comments_to_header(file_path:str, data_dict:dict="data_dict.json"):
    """
    Use this tool to add the data dictionary {data_dict.json} as comments to the header of an Excel file and save the output file.

    The function takes the Excel file path as argument and adds the {data_dict.json} as comments to each cell
    Start counting from column 0
    in the first row of the Excel file, using the following format:    
        * Column Number: <column_number>
        * Column Name: <column_name>
        * Data Type: <data_type>
        * Description: <description>

    Parameters
    ----------
    * file_path : str
        The path to the Excel file to be processed
    * data_dict : dict
        The data dictionary containing the column number, column name, data type, description, and number of null values

    Returns
    -------
    None
    """
    
    # Load the data dictionary
    data_dict = json.load(open(data_dict))

    # Load the workbook
    wb = load_workbook(file_path)

    # Get the active worksheet
    ws = wb.active

    # Iterate over each column in the first row (header)
    for n, col in enumerate(ws.iter_cols(min_row=1, max_row=1)):
        for header_cell in col:
            header_cell.comment = Comment(dedent(f"""\
                              ColName: {data_dict[str(n)]['ColName']}, 
                              DataType: {data_dict[str(n)]['DataType']},
                              Description: {data_dict[str(n)]['Description']}\
    """),'AI Agent')

    # Save the workbook
    st.write("Saving File... :floppy_disk:")
    wb.save('output.xlsx')

    # Create a download button
    with open('output.xlsx', 'rb') as f:
        st.download_button(
            label="Download output.xlsx",
            data=f,
            file_name='output.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

# Create the agent
def create_agent(apy_key):
    agent = Agent(
        model=Gemini(id="gemini-2.0-flash", api_key=apy_key),
        description= dedent("""\
                            You are an agent that reads the temp.csv dataset presented to you and 
                            based on the name and data type of each column header, determine the following information:
                            - The data types of each column
                            - The description of each column
                            - The first column numer is 0

                            Using the FileTools provided, create a data dictionary in JSON format that includes the below information:
                            {<ColNumber>: {ColName: <ColName>, DataType: <DataType>, Description: <Description>}}

                            If you are unable to determine the data type or description of a column, return 'N/A' for that column for the missing values.
                            \
                            """),
        tools=[ FileTools(read_files=True, save_files=True) ],
        retries=2,
        show_tool_calls=True
        )

    return agent


# Run the agent
if __name__ == "__main__":
    
    # Config page Streamlit
    st.set_page_config(layout="centered", 
                       page_title="Data Docs", 
                       page_icon=":paperclip:",
                       initial_sidebar_state="expanded")
    
    # Title
    st.title("Data Docs :paperclip:")
    st.subheader("Generate a data dictionary for your Excel file.")
    st.caption("1. Enter your Gemini API key and the path of the Excel file on the sidebar.")
    st.caption("2. Run the agent.")
    st.caption("3. The agent will generate a data dictionary and add it as comments to the header of the Excel file.")
    st.caption("ColName: <ColName> | DataType: <DataType> | Description: <Description>")
    
    # progress bar
    progress_bar = st.empty()
    progress_bar.progress(0)
    time.sleep(1)
    progress_bar.progress(10)
    time.sleep(3)
    progress_bar.progress(35)

    st.divider()


    with st.sidebar:
        # Enter your API key
        st.caption("Enter your API key and the path of the Excel file.")
        api_key = st.text_input("API key: ", placeholder="API key", type="password")
        
        # Upload file
        input_file = st.file_uploader("File upload", 
                                       type='xlsx')
        

        # Run the agent
        agent_run = st.button("Run")

        st.divider()

        # Reset session state
        if st.button("Reset Session"):
            st.session_state.clear()
            st.rerun()

    # Create the agent
    if agent_run:
        # Convert Excel file to CSV
        convert_to_csv(input_file)

        # Create the agent
        agent = create_agent(api_key)

        # Start the script
        st.write("Running Agent... :runner:")

        # Run the agent    
        agent.print_response(dedent(f"""\
                                1. Use FileTools to read the temp.csv as input to create the data dictionary for the columns in the dataset. 
                                2. Using the FileTools tool, save the data dictionary to a file named 'data_dict.json'.
                                \
                                """),
                        markdown=True)
    
        # Print the data dictionary
        st.write("Generating Data Dictionary... :page_facing_up:")
        with open('data_dict.json', 'r') as f:
            data_dict = json.load(f)
            st.json(data_dict, expanded=False)

        # Add comments to header
        add_comments_to_header(input_file, 'data_dict.json')

        # Remove temporary files
        st.write("Removing temporary files... :wastebasket:")
        os.remove('temp.csv')
        os.remove('data_dict.json')    
    
    # If file exists, show success message
    if os.path.exists('output.xlsx'):
        st.success("Done! :white_check_mark:")
        os.remove('output.xlsx')

    # Progress bar end
    progress_bar.progress(100)